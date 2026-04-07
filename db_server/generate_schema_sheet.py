"""
Connects to MySQL and generates a formatted Excel spreadsheet
showing every table and column in the Smart Factory databases.
"""

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '1234',
    'use_pure': True
}

DATABASES = ['sf_order', 'sf_inventory', 'sf_production', 'sf_report']

# ── Colours ──────────────────────────────────────────────────────────────────
C_DB_HEADER   = '1E1B4B'   # deep indigo  – database title row
C_TBL_HEADER  = '1E3A5F'   # dark blue    – table name row
C_COL_HEADER  = '1A1D27'   # near-black   – column-header row
C_ROW_ODD     = '0F1117'   # very dark    – odd data rows
C_ROW_EVEN    = '14171F'   # slightly lighter – even data rows
C_KEY_ACCENT  = '4ADE80'   # green        – PRI key text
C_IDX_ACCENT  = '60A5FA'   # blue         – MUL/UNI key text
C_TEXT        = 'E2E8F0'
C_TEXT_DIM    = '94A3B8'
C_TEXT_HEADER = 'A5B4FC'

def fill(hex_col):
    return PatternFill('solid', fgColor=hex_col)

def font(hex_col=C_TEXT, bold=False, sz=10):
    return Font(color=hex_col, bold=bold, size=sz, name='Consolas')

def border():
    s = Side(style='thin', color='2D3148')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=False)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=False)

def apply(cell, fill_=None, font_=None, align=None, border_=True):
    if fill_:   cell.fill    = fill_
    if font_:   cell.font    = font_
    if align:   cell.alignment = align
    if border_: cell.border  = border()

# ── Fetch schema from information_schema ─────────────────────────────────────
def get_schema(conn, db):
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT TABLE_NAME, TABLE_ROWS, TABLE_COMMENT
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = %s AND TABLE_TYPE = 'BASE TABLE'
        ORDER BY TABLE_NAME
    """, (db,))
    tables = cur.fetchall()

    schema = []
    for tbl in tables:
        tbl_name = tbl['TABLE_NAME']

        cur.execute("""
            SELECT
                COLUMN_NAME,
                ORDINAL_POSITION,
                COLUMN_TYPE,
                IS_NULLABLE,
                COLUMN_DEFAULT,
                COLUMN_KEY,
                EXTRA,
                COLUMN_COMMENT
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            ORDER BY ORDINAL_POSITION
        """, (db, tbl_name))
        cols = cur.fetchall()

        cur.execute("""
            SELECT INDEX_NAME, COLUMN_NAME, NON_UNIQUE
            FROM information_schema.STATISTICS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            ORDER BY INDEX_NAME, SEQ_IN_INDEX
        """, (db, tbl_name))
        idx_rows = cur.fetchall()

        # Group index columns
        indexes = {}
        for r in idx_rows:
            name = r['INDEX_NAME']
            indexes.setdefault(name, {'columns': [], 'unique': not r['NON_UNIQUE']})
            indexes[name]['columns'].append(r['COLUMN_NAME'])

        schema.append({'table': tbl_name, 'columns': cols, 'indexes': indexes})

    cur.close()
    return schema

# ── Write one sheet per database ─────────────────────────────────────────────
COL_HEADERS = ['#', 'Column', 'Type', 'Nullable', 'Default', 'Key', 'Extra', 'Comment']
COL_WIDTHS  = [4,   24,       26,     9,          18,        6,     20,      30       ]

def write_sheet(ws, db_name, schema):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '818CF8'

    row = 1

    # ── Database title ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_HEADERS))
    cell = ws.cell(row=row, column=1, value=f'  DATABASE: {db_name.upper()}')
    apply(cell, fill_=fill(C_DB_HEADER),
          font_=font(C_TEXT_HEADER, bold=True, sz=12), align=left())
    ws.row_dimensions[row].height = 26
    row += 1

    for tbl_info in schema:
        tbl_name = tbl_info['table']
        cols     = tbl_info['columns']
        indexes  = tbl_info['indexes']

        # ── Table name row ────────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_HEADERS))
        cell = ws.cell(row=row, column=1, value=f'  {tbl_name}')
        apply(cell, fill_=fill(C_TBL_HEADER),
              font_=font(C_TEXT_HEADER, bold=True, sz=11), align=left())
        ws.row_dimensions[row].height = 22
        row += 1

        # ── Column headers ────────────────────────────────────────────────────
        for ci, hdr in enumerate(COL_HEADERS, 1):
            cell = ws.cell(row=row, column=ci, value=hdr)
            apply(cell, fill_=fill(C_COL_HEADER),
                  font_=font(C_TEXT_DIM, bold=True), align=center())
        ws.row_dimensions[row].height = 18
        row += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        for i, col in enumerate(cols):
            bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
            key = col['COLUMN_KEY'] or ''

            key_color = C_TEXT
            if key == 'PRI':
                key_color = C_KEY_ACCENT
            elif key in ('MUL', 'UNI'):
                key_color = C_IDX_ACCENT

            values = [
                col['ORDINAL_POSITION'],
                col['COLUMN_NAME'],
                col['COLUMN_TYPE'],
                col['IS_NULLABLE'],
                str(col['COLUMN_DEFAULT']) if col['COLUMN_DEFAULT'] is not None else '',
                key,
                col['EXTRA'] or '',
                col['COLUMN_COMMENT'] or '',
            ]

            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                fc = key_color if ci in (1, 2, 6) else C_TEXT_DIM if ci in (4, 5, 7) else C_TEXT
                apply(cell, fill_=fill(bg), font_=font(fc),
                      align=center() if ci in (1, 4, 6) else left())
            ws.row_dimensions[row].height = 16
            row += 1

        # ── Indexes block ─────────────────────────────────────────────────────
        if indexes:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_HEADERS))
            cell = ws.cell(row=row, column=1, value='  Indexes')
            apply(cell, fill_=fill('0F1117'), font_=font('475569', bold=True), align=left())
            ws.row_dimensions[row].height = 14
            row += 1

            for idx_name, idx_info in indexes.items():
                unique_label = 'UNIQUE' if idx_info['unique'] else ''
                label = f"  {idx_name}  ({', '.join(idx_info['columns'])})"
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_HEADERS) - 1)
                cell = ws.cell(row=row, column=1, value=label)
                apply(cell, fill_=fill('0A0D14'), font_=font('475569'), align=left())
                cell2 = ws.cell(row=row, column=len(COL_HEADERS), value=unique_label)
                apply(cell2, fill_=fill('0A0D14'), font_=font(C_IDX_ACCENT), align=center())
                ws.row_dimensions[row].height = 14
                row += 1

        row += 1  # blank spacer between tables

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    conn = mysql.connector.connect(**DB_CONFIG)
    wb = Workbook()
    wb.remove(wb.active)

    for db in DATABASES:
        schema = get_schema(conn, db)
        ws = wb.create_sheet(title=db)
        write_sheet(ws, db, schema)
        print(f'  Written: {db}  ({len(schema)} tables)')

    conn.close()
    out = 'schema_overview.xlsx'
    wb.save(out)
    print(f'\nSaved → {out}')

if __name__ == '__main__':
    main()
